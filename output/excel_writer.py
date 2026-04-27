import os

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

SEVERITY_ORDER = {
    "critical": 0,
    "high": 1,
    "medium": 2,
    "low": 3,
    "info": 4,
    "": 5,
}

SEVERITY_FILL = {
    "critical": "FFC7CE",
    "high": "FCE4D6",
    "medium": "FFF2CC",
    "low": "E2F0D9",
    "info": "D9EAF7",
}

HEADER_FILL = PatternFill("solid", fgColor="1F2937")
HEADER_FONT = Font(color="FFFFFF", bold=True)


def _severity_key(value):
    return SEVERITY_ORDER.get(str(value).lower(), 5)


def _normalize_severity(value):
    value = str(value or "").strip().lower()
    if value.isdigit():
        mapping = {"4": "critical", "3": "high", "2": "medium", "1": "low", "0": "info"}
        return mapping.get(value, value)
    return value


def _safe_join(values):
    if isinstance(values, str):
        return values
    if isinstance(values, (list, tuple, set)):
        return ", ".join(str(v) for v in values if v)
    return str(values or "")


def build_summary_df(findings):
    rows = {}

    for finding in findings:
        vuln = finding.get("category") or finding.get("issue") or "unknown"
        key = str(vuln).strip() or "unknown"
        rows[key] = rows.get(key, 0) + 1

    summary_rows = [
        {"Vulnerability": vuln, "Count": count}
        for vuln, count in sorted(rows.items(), key=lambda item: (-item[1], item[0]))
    ]

    return pd.DataFrame(summary_rows, columns=["Vulnerability", "Count"])


def build_findings_df(findings):
    rows = []
    vuln_counts = {}

    for finding in findings:
        vuln = finding.get("category") or finding.get("issue") or "unknown"
        key = str(vuln).strip() or "unknown"
        vuln_counts[key] = vuln_counts.get(key, 0) + 1

    for finding in findings:
        severity = _normalize_severity(finding.get("severity", ""))
        vuln = finding.get("category") or finding.get("issue") or "unknown"
        vuln_key = str(vuln).strip() or "unknown"

        rows.append({
            "Severity": severity,
            "Vulnerability": vuln,
            "Count": vuln_counts.get(vuln_key, 1),
            "Target": finding.get("target", ""),
            "Tools": finding.get("tools", finding.get("tool", "")),
            "Issue": finding.get("issue", ""),
            "CVE": _safe_join(finding.get("cve", "")),
            "Description": finding.get("description", ""),
            "Recommendation": finding.get("recommendation", ""),
        })

    df = pd.DataFrame(
        rows,
        columns=[
            "Severity",
            "Vulnerability",
            "Count",
            "Target",
            "Tools",
            "Issue",
            "CVE",
            "Description",
            "Recommendation",
        ],
    )

    if df.empty:
        return df

    df["SeverityRank"] = df["Severity"].map(_severity_key)
    df = df.sort_values(by=["SeverityRank", "Vulnerability", "Target"], ascending=[True, True, True])
    df = df.drop(columns=["SeverityRank"])

    return df


def style_excel_report(output_file):
    wb = load_workbook(output_file)

    for ws in wb.worksheets:
        ws.freeze_panes = "A2"
        ws.sheet_view.showGridLines = False

        for cell in ws[1]:
            cell.fill = HEADER_FILL
            cell.font = HEADER_FONT
            cell.alignment = Alignment(horizontal="center", vertical="center")

        if ws.title == "Summary":
            for row in ws.iter_rows(min_row=2, min_col=1, max_col=2):
                for cell in row:
                    cell.alignment = Alignment(vertical="top")
        elif ws.title == "Findings":
            severity_col = None
            vulnerability_col = None
            for idx, cell in enumerate(ws[1], start=1):
                if cell.value == "Severity":
                    severity_col = idx
                elif cell.value == "Vulnerability":
                    vulnerability_col = idx

            for row in ws.iter_rows(min_row=2):
                severity_value = ""
                if severity_col:
                    severity_value = str(row[severity_col - 1].value or "").lower()

                fill_color = SEVERITY_FILL.get(severity_value)
                if fill_color:
                    for cell in row:
                        cell.fill = PatternFill("solid", fgColor=fill_color)
                        cell.alignment = Alignment(vertical="top", wrap_text=True)
                else:
                    for cell in row:
                        cell.alignment = Alignment(vertical="top", wrap_text=True)

                if vulnerability_col:
                    row[vulnerability_col - 1].font = Font(bold=True)

        for column_cells in ws.columns:
            values = []
            col_idx = column_cells[0].column
            for cell in column_cells:
                if cell.value is not None:
                    values.append(str(cell.value))
            width = max([len(v) for v in values] + [len(str(column_cells[0].value or ""))])
            ws.column_dimensions[get_column_letter(col_idx)].width = min(max(width + 2, 12), 60)

    wb.save(output_file)


def export_report(findings, output_file):
    os.makedirs(os.path.dirname(output_file) or ".", exist_ok=True)

    summary_df = build_summary_df(findings)
    findings_df = build_findings_df(findings)

    with pd.ExcelWriter(output_file, engine="openpyxl") as writer:
        summary_df.to_excel(writer, sheet_name="Summary", index=False)
        findings_df.to_excel(writer, sheet_name="Findings", index=False)

    style_excel_report(output_file)
